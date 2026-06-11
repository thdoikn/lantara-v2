"""
Analytics API — Phase 2.

Endpoints (all require is_staff):
  GET /api/v1/analytics/summary/          — overall counts + SLA stats
  GET /api/v1/analytics/by-sektor/        — per-sektor breakdown
  GET /api/v1/analytics/sla/             — SLA breach/at-risk counts
  GET /api/v1/analytics/trend/?days=30   — daily submission counts (last N days)
  GET /api/v1/analytics/export/excel/    — Excel workbook download
"""
import io
from datetime import date, timedelta

from django.db.models import Case, Count, F, IntegerField, Value, When
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.submissions.models import Submission


class SummaryView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        qs = Submission.objects.all()
        total = qs.count()
        by_status = dict(
            qs.values_list("status").annotate(c=Count("id")).values_list("status", "c")
        )
        active = qs.filter(
            status__in=["submitted", "in_review", "revision", "publishing", "collection"]
        ).count()
        approved = qs.filter(status="approved").count()
        rejected = qs.filter(status="rejected").count()
        sla_breached = qs.filter(is_sla_breached=True).count()
        sla_at_risk = qs.filter(is_sla_at_risk=True, is_sla_breached=False).count()

        return Response(
            {
                "total": total,
                "active": active,
                "approved": approved,
                "rejected": rejected,
                "sla_breached": sla_breached,
                "sla_at_risk": sla_at_risk,
                "by_status": by_status,
            }
        )


class BySektorView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        rows = (
            Submission.objects.values(
                sektor=F("permit_type__sektor__name"),
                sektor_key=F("permit_type__sektor__key"),
            )
            .annotate(
                total=Count("id"),
                active=Count(
                    Case(
                        When(
                            status__in=[
                                "submitted",
                                "in_review",
                                "revision",
                                "publishing",
                                "collection",
                            ],
                            then=Value(1),
                        ),
                        output_field=IntegerField(),
                    )
                ),
                approved=Count(Case(When(status="approved", then=Value(1)), output_field=IntegerField())),
                rejected=Count(Case(When(status="rejected", then=Value(1)), output_field=IntegerField())),
                breached=Count(Case(When(is_sla_breached=True, then=Value(1)), output_field=IntegerField())),
            )
            .order_by("sektor")
        )
        return Response(list(rows))


class SLAView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        active_qs = Submission.objects.filter(
            status__in=["submitted", "in_review", "revision", "publishing", "collection"]
        )
        breached = active_qs.filter(is_sla_breached=True).count()
        at_risk = active_qs.filter(is_sla_at_risk=True, is_sla_breached=False).count()
        on_time = active_qs.filter(is_sla_breached=False, is_sla_at_risk=False).count()

        return Response(
            {
                "active_total": breached + at_risk + on_time,
                "breached": breached,
                "at_risk": at_risk,
                "on_time": on_time,
            }
        )


class TrendView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        days = min(int(request.query_params.get("days", 30)), 365)
        since = timezone.now() - timedelta(days=days)

        rows = (
            Submission.objects.filter(created_at__gte=since)
            .annotate(day=TruncDate("created_at"))
            .values("day")
            .annotate(count=Count("id"))
            .order_by("day")
        )

        # Fill gaps with 0
        data: dict[date, int] = {r["day"]: r["count"] for r in rows}
        result = []
        for i in range(days):
            d = (timezone.now() - timedelta(days=days - 1 - i)).date()
            result.append({"date": d.isoformat(), "count": data.get(d, 0)})

        return Response(result)


class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return Response(
                {"detail": "openpyxl not installed"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()

        # ── Sheet 1: All submissions ──────────────────────────────────────────
        ws = wb.active
        ws.title = "Permohonan"

        headers = [
            "No. Referensi",
            "Jenis Izin",
            "Sektor",
            "Pemohon",
            "Email",
            "Status",
            "Tanggal Ajukan",
            "SLA Berakhir",
            "SLA Terlampaui",
        ]
        header_fill = PatternFill("solid", fgColor="428A40")
        header_font = Font(bold=True, color="FFFFFF")
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        qs = (
            Submission.objects.select_related("permit_type__sektor")
            .order_by("-created_at")
        )
        for row_idx, sub in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=sub.reference_number)
            ws.cell(row=row_idx, column=2, value=sub.permit_type.name if sub.permit_type else "")
            ws.cell(
                row=row_idx,
                column=3,
                value=sub.permit_type.sektor.name if sub.permit_type and sub.permit_type.sektor else "",
            )
            ws.cell(row=row_idx, column=4, value=sub.applicant.get_full_name() if sub.applicant else "")
            ws.cell(row=row_idx, column=5, value=sub.applicant.email if sub.applicant else "")
            ws.cell(row=row_idx, column=6, value=sub.get_status_display())
            ws.cell(
                row=row_idx,
                column=7,
                value=sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else "",
            )
            ws.cell(
                row=row_idx,
                column=8,
                value=sub.sla_due_at.strftime("%Y-%m-%d") if sub.sla_due_at else "",
            )
            ws.cell(row=row_idx, column=9, value="Ya" if sub.is_sla_breached else "Tidak")

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Sheet 2: By-sektor summary ────────────────────────────────────────
        ws2 = wb.create_sheet("Rekap per Sektor")
        s2_headers = ["Sektor", "Total", "Aktif", "Disetujui", "Ditolak", "SLA Terlampaui"]
        for col, hdr in enumerate(s2_headers, 1):
            cell = ws2.cell(row=1, column=col, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        rows = (
            Submission.objects.values(
                sektor=F("permit_type__sektor__name"),
            )
            .annotate(
                total=Count("id"),
                active=Count(
                    Case(
                        When(
                            status__in=["submitted", "in_review", "revision", "publishing", "collection"],
                            then=Value(1),
                        ),
                        output_field=IntegerField(),
                    )
                ),
                approved=Count(Case(When(status="approved", then=Value(1)), output_field=IntegerField())),
                rejected=Count(Case(When(status="rejected", then=Value(1)), output_field=IntegerField())),
                breached=Count(Case(When(is_sla_breached=True, then=Value(1)), output_field=IntegerField())),
            )
            .order_by("sektor")
        )
        for row_idx, r in enumerate(rows, 2):
            ws2.cell(row=row_idx, column=1, value=r["sektor"] or "—")
            ws2.cell(row=row_idx, column=2, value=r["total"])
            ws2.cell(row=row_idx, column=3, value=r["active"])
            ws2.cell(row=row_idx, column=4, value=r["approved"])
            ws2.cell(row=row_idx, column=5, value=r["rejected"])
            ws2.cell(row=row_idx, column=6, value=r["breached"])

        # ─────────────────────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        today = date.today().isoformat()
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="lantara-export-{today}.xlsx"'
        return response
